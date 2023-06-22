from io import BytesIO
import openpyxl
from openpyxl.styles import Font
from django.shortcuts import render, redirect
from .models import Member
from .forms import LoginForm
from django.contrib.auth import authenticate, logout, login as auth_login
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse


@login_required
def select_zone(request):
    return render(request, 'members/select_zone.html', {})


def login(request):
    context = {}

    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            user_name = form.cleaned_data['user_name']
            user_password = form.cleaned_data['user_password']

            user = authenticate(username=user_name, password=user_password)

            if user is not None:
                print("Auth succeeded")
                auth_login(request, user)
                return redirect('home')
            else:
                context['form'] = form
                context['error_message'] = 'Username/Password is incorrect'
                print("Auth failed")
        else:
            context['form'] = form
            context['error_message'] = 'Please correct the errors below.'
            print("Form is invalid")
    else:
        form = LoginForm()
        context['form'] = form

    return render(request, 'members/login.html', context)


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def add_member(request):
    if request.method == 'POST':
        surname = request.POST.get('surname')
        other_name = request.POST.get('other_name')
        phone_number = request.POST.get('phone_number')
        date_of_birth = request.POST.get('date_of_birth')
        wedding_anniversary = request.POST.get('wedding_anniversary')
        missional_community = request.POST.get('missional_community')

        m = Member(Surname=surname, Other_name=other_name, Phone_number=phone_number, Wedding_anniversary=wedding_anniversary, Date_of_birth=date_of_birth, Missional_Community=missional_community)
        m.save()

        return redirect('home')
        
    return render(request, 'members/add_member.html', {})


@login_required
def manage_database(request):
    members = Member.objects.all()
    columns = [f.name for f in Member._meta.get_fields()]
    rows = [[getattr(member, col) for col in columns] for member in members]

    return render(request, 'members/manage_database.html', {'columns': columns, 'rows': rows})


@login_required
def delete_rows(request):
    if request.method == 'POST':
        selected_rows = request.POST['selected_rows'].split(',')
        Member.objects.filter(pk__in=selected_rows).delete()
        return redirect('manage_database')


@login_required
def home(request):
    return render(request, 'members/home.html', {})


@login_required
def update_member(request, member_id):
    member = Member.objects.get(pk=member_id)

    if request.method == 'POST':
        member.Surname = request.POST['surname']
        member.Other_name = request.POST['other_name']
        member.Phone_number = request.POST['phone_number']
        member.Date_of_birth = request.POST['date_of_birth']
        member.Wedding_anniversary = request.POST['wedding_anniversary']
        member.Missional_Community = request.POST['missional_community']

        member.save()

        return redirect('manage_database')

    context = {
        'member': member,
    }

    return render(request, 'members/update_member.html', context)


@login_required
def update_rows(request):
    if request.method == 'POST':
        member_id = request.POST['member_id']
        return redirect('update_member', member_id=member_id)


def export_excel(request):
    members = Member.objects.all()

    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define the font for bold text
    bold_font = Font(bold=True)

    # Create column headers with bold text
    headers = ['Surname', 'Other Name', 'Phone Number', 'Date of Birth', 'Wedding Anniversary', 'Missional Community']
    for col_num, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws['{}1'.format(col_letter)] = header
        ws['{}1'.format(col_letter)].font = bold_font

    # Add data to the worksheet
    for row_num, member in enumerate(members, 2):
        ws.cell(row=row_num, column=1, value=member.Surname)
        ws.cell(row=row_num, column=2, value=member.Other_name)
        ws.cell(row=row_num, column=3, value=member.Phone_number)
        ws.cell(row=row_num, column=4, value=member.Date_of_birth)
        ws.cell(row=row_num, column=5, value=member.Wedding_anniversary)
        ws.cell(row=row_num, column=6, value=member.Missional_Community)

    # Save the workbook to a bytes buffer
    buffer = BytesIO()
    wb.save(buffer)

    # Serve the xlsx file as a response
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=members.xlsx'
    return response

def ajax_logout(request):
    logout(request)
    return JsonResponse({"success": True})